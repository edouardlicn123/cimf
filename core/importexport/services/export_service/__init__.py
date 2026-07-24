"""
ExportService - 导出服务

提供通用的数据导出功能，支持 CSV/Excel 格式
支持自动发现模块字段，无需模块主动配置
"""

import contextlib
import logging

from django.http import HttpResponse
from django.utils.timezone import now as timezone_now
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .field_service import FieldServiceMixin
from .query_service import QueryServiceMixin
from .value_resolver import ValueResolverMixin

logger = logging.getLogger(__name__)


class ExportService(FieldServiceMixin, QueryServiceMixin, ValueResolverMixin):
    """数据导出服务"""

    FORMAT_CSV = "csv"
    FORMAT_XLSX = "xlsx"

    @classmethod
    def get_preview(
        cls, node_type_slug: str, field_names: list[str], filters: list[dict] | None = None, limit: int = 5
    ) -> list[dict]:
        """获取数据预览"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields_info = cls.get_fields_info(node_type_slug, field_names)
        field_type_map = {f["name"]: f["type"] for f in fields_info}

        preview_data = []
        for item in queryset[:limit]:
            row = cls._convert_to_row(item, field_names, field_type_map, node_type_slug)
            preview_data.append(row)

        return preview_data

    @classmethod
    def get_record_count(cls, node_type_slug: str, filters: list[dict] | None = None) -> int:
        """获取记录总数"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        return queryset.count()

    @classmethod
    def export(
        cls, node_type_slug: str, field_names: list[str], export_format: str = "csv", filters: list[dict] | None = None
    ) -> HttpResponse:
        """执行导出"""
        queryset = cls._get_filtered_queryset(node_type_slug, filters)
        fields_info = cls.get_fields_info(node_type_slug, field_names)
        field_type_map = {f["name"]: f["type"] for f in fields_info}

        rows = [cls._convert_to_row(item, field_names, field_type_map, node_type_slug) for item in queryset]

        filename = cls.generate_filename(node_type_slug, export_format)

        if export_format == cls.FORMAT_CSV:
            return cls._export_csv(rows, fields_info, filename)
        else:
            return cls._export_xlsx(rows, fields_info, filename)

    @classmethod
    def generate_filename(cls, node_type_slug: str, export_format: str) -> str:
        """生成导出文件名"""
        timestamp = timezone_now().strftime("%Y%m%d_%H%M%S")
        return f"{node_type_slug}_{timestamp}.{export_format}"

    @classmethod
    def _export_csv(cls, rows: list[dict], fields: list[dict], filename: str) -> HttpResponse:
        """导出为 CSV"""
        from core.utils.response import csv_response  # noqa: PLC0415

        headers = [f["label"] for f in fields]
        data_rows = [[row.get(f["name"], "") for f in fields] for row in rows]

        return csv_response(headers, data_rows, filename)

    @classmethod
    def _export_xlsx(cls, rows: list[dict], fields: list[dict], filename: str) -> HttpResponse:
        """导出为 XLSX"""

        wb = Workbook()
        ws = wb.active
        ws.title = "Data"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        headers = [f["label"] for f in fields]
        ws.append(headers)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

        for row in rows:
            ws.append(
                [
                    f"'{v}" if isinstance(v, str) and v.startswith(("=", "+", "-", "@")) else v
                    for f in fields
                    if (v := row.get(f["name"], "")) is not None
                ]
            )

        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            column = get_column_letter(i)
            for cell in col:
                with contextlib.suppress(Exception):
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        wb.save(response)  # noqa: CIMF_W006 — openpyxl Workbook.save()，非 Django model
        return response
